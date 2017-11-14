# encoding: utf-8
from random import randint
from openpyxl import *

wb = load_workbook('result.xlsx')
ws = wb.active

wb_target = Workbook()
ws_target = wb_target.create_sheet("Sheet1")
ws_target.append([u'序号', u'班级', u'学号', u'姓名', \
                  u'学制', u'入学时间', u'购票区间(家庭)', \
                  u'身份证号'])

row_num = 1
for row in ws.iter_rows(min_row=2):
    class_ = row[10].value
    if class_ == u'(空)':
        class_ = randint(2, 15)
    sid = row[8].value
    name = row[7].value
    years = 3
    start_year = '2017.09.03'
    dest = row[11].value.split('-')[1]
    id_ = row[12].value
    ws_target.append([row_num, class_, sid, name, years, start_year, \
            dest, id_])
    row_num += 1


print wb_target.get_sheet_names()
default_sheet = wb_target.get_sheet_by_name('Sheet')
wb_target.remove_sheet(default_sheet)
print wb_target.get_sheet_names()
wb_target.save('processed.xlsx')
